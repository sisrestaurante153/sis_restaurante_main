from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .normalize import normalize_header, normalize_name


FINAL_CODE_RE = re.compile(r"^(seg|ter2?|qua|qui|sex|sab|sab|sáb|frg|fre|lga|stf)\s+[pmg]$", re.I)


@dataclass
class SheetPreview:
    name: str
    kind: str
    product_name: str | None


def open_workbook(path: Path):
    return load_workbook(path, data_only=True, read_only=True)


def preview_rows(worksheet, limit_rows: int = 40, limit_cols: int = 14) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for row in worksheet.iter_rows(max_row=limit_rows, max_col=limit_cols, values_only=True):
        if any(value is not None and str(value).strip() for value in row):
            rows.append(list(row))
    return rows


def extract_labeled_value(rows: list[list[Any]], label: str) -> str | None:
    wanted = normalize_header(label)
    for row in rows:
        for index, value in enumerate(row[:-1]):
            if normalize_header(value) == wanted:
                next_value = row[index + 1]
                if next_value is None:
                    return None
                return str(next_value).strip()
    return None


def is_ficha_sheet(rows: list[list[Any]]) -> bool:
    flattened = [normalize_header(value) for row in rows for value in row if value is not None]
    has_title = "ficha tecnica" in flattened
    has_product = "produto" in flattened
    has_ingredients = "ingredientes" in flattened
    return has_title and has_product and has_ingredients


def looks_like_final_sheet(sheet_name: str, product_name: str | None) -> bool:
    normalized_sheet = normalize_name(sheet_name)
    normalized_product = normalize_name(product_name)

    if FINAL_CODE_RE.match(sheet_name.strip()):
        return True

    patterns = ["prato ", "especial ", "marmita ", "combo ", "salada "]
    return any(
        normalized_sheet.startswith(pattern) or normalized_product.startswith(pattern)
        for pattern in patterns
    )


def classify_sheet(sheet_name: str, rows: list[list[Any]]) -> SheetPreview:
    product_name = extract_labeled_value(rows, "Produto")

    if sheet_name == "TABELA VMARKET":
        return SheetPreview(sheet_name, "vmarket", product_name)
    if sheet_name == "EMBALAGENS":
        return SheetPreview(sheet_name, "embalagens", product_name)
    if sheet_name == "PESOS":
        return SheetPreview(sheet_name, "pesos", product_name)
    if sheet_name == "MODELO":
        return SheetPreview(sheet_name, "ignored", product_name)
    if is_ficha_sheet(rows):
        kind = "final_composition" if looks_like_final_sheet(sheet_name, product_name) else "recipe"
        return SheetPreview(sheet_name, kind, product_name)
    return SheetPreview(sheet_name, "unknown", product_name)


def inspect_workbook(path: Path) -> dict[str, Any]:
    workbook = open_workbook(path)
    previews: list[SheetPreview] = []

    for sheet_name in workbook.sheetnames:
      rows = preview_rows(workbook[sheet_name])
      previews.append(classify_sheet(sheet_name, rows))

    summary = {
        "vmarket_sheets": sum(sheet.kind == "vmarket" for sheet in previews),
        "embalagens_sheets": sum(sheet.kind == "embalagens" for sheet in previews),
        "pesos_sheets": sum(sheet.kind == "pesos" for sheet in previews),
        "recipe_sheet_candidates": sum(sheet.kind == "recipe" for sheet in previews),
        "final_sheet_candidates": sum(sheet.kind == "final_composition" for sheet in previews),
        "unknown_sheets": sum(sheet.kind == "unknown" for sheet in previews),
    }

    return {
        "summary": summary,
        "sheets": [
            {
                "name": sheet.name,
                "kind": sheet.kind,
                "product_name": sheet.product_name,
            }
            for sheet in previews
        ],
    }
